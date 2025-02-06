import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
import os


class MetricsSaver:
    def __init__(self, filename=None):
        self.filename = filename

    def save_metrics(self, metrics, filename=None, sheet_name='Risultati Metriche'):
        if filename is None:
            filename = self.filename

        if not filename:
            filename = "risultati_metriche.xlsx"

        self.filename = filename
        abs_path = os.path.join(os.getcwd(), self.filename)

        metric_values = []
        for metric_name, metric_list in metrics.items():
            if isinstance(metric_list, list):
                str_values = ', '.join(map(str, metric_list))
            else:
                str_values = str(metric_list)
            metric_values.append((metric_name, str_values))

        df = pd.DataFrame(metric_values, columns=['Metriche', 'Valori'])

        try:
            mode = 'a' if os.path.exists(abs_path) else 'w'
            with pd.ExcelWriter(abs_path, engine='openpyxl', mode=mode) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"Errore durante il salvataggio delle metriche su Excel: {e}")
        else:
            print(f"Le metriche sono state salvate nel foglio '{sheet_name}' del file Excel:\n  {abs_path}")

    def save_plot_to_excel(self, temp_file, sheet_name):
        abs_path = os.path.join(os.getcwd(), self.filename)

        try:
            wb = load_workbook(abs_path)
            try:
                ws = wb[sheet_name]
            except KeyError:
                ws = wb.create_sheet(title=sheet_name)

            img = openpyxl.drawing.image.Image(temp_file)
            img.width = img.width * 0.8
            img.height = img.height * 0.8
            ws.add_image(img, 'A1')

            wb.save(abs_path)

            try:
                os.remove(temp_file)
            except OSError:
                print(f"Impossibile eliminare il file temporaneo: {temp_file}")
        except Exception as e:
            print(f"Errore durante il salvataggio del plot su Excel: {e}")

    def metrics_plot(self, metrics):
        labels = list(metrics.keys())
        metric_scores = [vals if isinstance(vals, list) else [vals] for vals in metrics.values()]

        plt.figure(figsize=(12, 6))
        plt.boxplot(metric_scores, tick_labels=labels)
        plt.xlabel('Metriche')
        plt.ylabel('Valori')
        plt.title('Metriche di validazione (Boxplot)')
        plt.grid(True)

        temp_file1 = 'temp_plot_boxplot.png'
        plt.savefig(temp_file1)
        self.save_plot_to_excel(temp_file1, 'Plot Andamento Metriche')

        num_experiments = len(metric_scores[0])
        x_experiments = range(1, num_experiments + 1)

        plt.figure(figsize=(12, 6))
        for label, values in zip(labels, metric_scores):
            plt.plot(x_experiments, values, marker='o', label=label)

        plt.xlabel('Esperimenti (K)')
        plt.ylabel('Valore della metrica')
        plt.title('Andamento delle metriche nei vari esperimenti')
        plt.legend()
        plt.grid(True)

        temp_file2 = 'temp_plot_line.png'
        plt.savefig(temp_file2)
        self.save_plot_to_excel(temp_file2, 'Andamento Metriche in K')

        plt.show()

    def plot_confusion_matrix(self, y_true, y_pred):
        unique_labels = [2, 4]
        cm = np.zeros((len(unique_labels), len(unique_labels)), dtype=int)
        label_to_index = {label: idx for idx, label in enumerate(unique_labels)}

        for true, pred in zip(y_true, y_pred):
            cm[label_to_index[true], label_to_index[pred]] += 1

        plt.figure(figsize=(8, 6))
        plt.imshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
        plt.title('Confusion Matrix')
        plt.colorbar()
        tick_marks = np.arange(len(unique_labels))
        plt.xticks(tick_marks, unique_labels, rotation=45)
        plt.yticks(tick_marks, unique_labels)

        fmt = 'd'
        thresh = cm.max() / 2.
        for i, j in np.ndindex(cm.shape):
            plt.text(j, i, format(cm[i, j], fmt),
                     ha="center", va="center",
                     color="white" if cm[i, j] > thresh else "black")

        plt.ylabel('True Label')
        plt.xlabel('Predicted Label')
        plt.tight_layout()

        temp_file = 'temp_confusion_matrix.png'
        plt.savefig(temp_file)
        self.save_plot_to_excel(temp_file, 'Confusion Matrix')

        plt.show()

    def plot_roc_curve(self, y_test: np.ndarray, y_score: np.ndarray, auc_label: str = "ROC Curve"):
        """
        Plotta la curva ROC per un esperimento e mostra l'immagine.

        param y_test: Array dei valori reali.
        param y_score: Array dei punteggi continui previsti.
        param auc_label: Etichetta da mostrare nel grafico (con il valore dell'AUC incluso).
        """

        # Ordina i dati in base ai punteggi
        y_test = np.asarray(y_test)
        y_score = np.asarray(y_score)
        sorted_indices = np.argsort(y_score)
        y_true_sorted = y_test[sorted_indices]

        # Calcola TPR e FPR
        total_positives = np.sum(y_true_sorted == 4)  # Classe positiva come esempio
        total_negatives = np.sum(y_true_sorted != 4)

        TPR = np.cumsum(y_true_sorted == 4) / total_positives
        FPR = np.cumsum(y_true_sorted != 4) / total_negatives

        # Plot della curva ROC
        plt.figure(figsize=(8, 6))
        plt.plot(TPR, FPR, color='blue', lw=2, label=auc_label)
        plt.plot([0, 1], [0, 1], color='gray', linestyle='--', lw=2, label='Random Classifier')
        plt.xlabel('False Positive Rate (FPR)')
        plt.ylabel('True Positive Rate (TPR)')
        plt.title('ROC Curve')
        plt.legend(loc='lower right')

        # Mostra il grafico
        temp_file = 'temp_roc_curve.png'
        plt.savefig(temp_file)
        self.save_plot_to_excel(temp_file, 'ROC Curve')

        plt.show()

    def salva_risultati_excel(results, metrics_by_experiment):
        """
        Salva i risultati delle metriche in un file Excel e genera i plot della confusion matrix e della ROC curve.

        :param results: Lista di tuple (y_test, predictions, predicted_proba).
        :param metrics_by_experiment: Dizionario delle metriche calcolate per ciascun esperimento.
        """
        # 1) Chiedi all'utente dove salvare il file Excel
        excel_path = input(
            "\nInserisci il percorso completo per il file Excel dove salvare (es. C:/risultati/risultati_metriche.xlsx): ").strip()
        if not excel_path.lower().endswith(".xlsx"):
            excel_path += ".xlsx"

        # 2) Crea un'istanza di MetricsSaver con il percorso scelto
        visual_metrics = MetricsSaver(filename=excel_path)

        # 3) Salva su Excel tutte le metriche (le liste di K esperimenti per ogni metrica)
        visual_metrics.save_metrics(metrics_by_experiment, sheet_name='Risultati Metriche')

        # 4) Mostra e salva i plot (boxplot e line plot) delle metriche su Excel
        visual_metrics.metrics_plot(metrics_by_experiment)

        # 5) (Facoltativo) Plot Confusion Matrix per ciascun esperimento
        plot_cm = input(
            "\nVuoi plottare e salvare la Confusion Matrix per ognuno dei K esperimenti? (s/n): ").strip().lower()
        if plot_cm == 's':
            for i, (y_test, predictions, predicted_proba) in enumerate(results):
                print(f"Plot Confusion Matrix - Esperimento {i + 1}")
                visual_metrics.plot_confusion_matrix(y_test, predictions)

        # Prepara i risultati per il plot ROC
        new_results = [(y_test, predicted_proba) for y_test, _, predicted_proba in results]

        # 6) (Facoltativo) Plot ROC per ciascun esperimento
        plot_roc = input("\nVuoi plottare e salvare la ROC Curve per ognuno dei K esperimenti? (s/n): ").strip().lower()
        if plot_roc == 's':
            for i, (y_test, predicted_proba) in enumerate(new_results):
                print(f"Plot ROC Curve - Esperimento {i + 1}")
                visual_metrics.plot_roc_curve(y_test, predicted_proba)

        print("\nFine esecuzione. Risultati salvati in:", excel_path)

