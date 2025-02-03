import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
import os

class MetricsSaver:
    def __init__(self, filename=None):
        self.filename = filename

    def save_metrics(self, metrics, filename=None, sheet_name='Risultati Metriche'):
        """
        Salva le metriche su file Excel.

        Parameters
        ----------
        metrics : dict
            Dizionario delle metriche, dove ogni metrica è una lista di valori
            (uno per ogni esperimento).
            Esempio:
                {
                    "Accuracy": [0.8, 0.85, 0.9],
                    "Error Rate": [0.2, 0.15, 0.1],
                    ...
                }

        filename : str, optional
            Nome del file Excel in cui salvare. Se non fornito,
            userà quello di self.filename.

        sheet_name : str, optional
            Nome del foglio in cui salvare le metriche.
        """
        if filename is None:
            filename = self.filename

        if not filename:
            # Se non c'è un filename, usiamo un default
            filename = "risultati_metriche.xlsx"
    
        self.filename = filename

        # Non creiamo percorsi: ci limitiamo alla directory corrente
        # Per sicurezza, se la cartella corrente fosse cambiata, ricava il path completo
        abs_path = os.path.join(os.getcwd(), self.filename)

        # Prepara i dati per il DataFrame
        metric_values = []
        for metric_name, metric_list in metrics.items():
            if isinstance(metric_list, list):
                str_values = ', '.join(map(str, metric_list))
            else:
                str_values = str(metric_list)
            metric_values.append((metric_name, str_values))

        df = pd.DataFrame(metric_values, columns=['Metriche', 'Valori'])

        # Scrittura su Excel: se il file esiste, usa append ('a'), altrimenti crea ('w')
        try:
            mode = 'a' if os.path.exists(abs_path) else 'w'
            with pd.ExcelWriter(abs_path, engine='openpyxl', mode=mode) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"Errore durante il salvataggio delle metriche su Excel: {e}")
        else:
            print(f"Le metriche sono state salvate nel foglio '{sheet_name}' del file Excel:\n  {abs_path}")

    def save_plot_to_excel(self, temp_file, sheet_name):
        """
        Salva un plot in un foglio Excel esistente o nuovo.

        Parameters
        ----------
        temp_file : str
            Nome dell'immagine temporanea salvata dal plot.
        sheet_name : str
            Nome del foglio in cui salvare il plot.
        """
        abs_path = os.path.join(os.getcwd(), self.filename)

        try:
            wb = load_workbook(abs_path)
            try:
                ws = wb[sheet_name]
            except KeyError:
                ws = wb.create_sheet(title=sheet_name)

            # Inserisce l'immagine
            img = openpyxl.drawing.image.Image(temp_file)
            # (Opzionale) riduci dimensioni
            img.width = img.width * 0.8
            img.height = img.height * 0.8
            ws.add_image(img, 'A1')

            wb.save(abs_path)

            # Elimina il file temporaneo
            try:
                os.remove(temp_file)
            except OSError:
                print(f"Impossibile eliminare il file temporaneo: {temp_file}")
        except Exception as e:
            print(f"Errore durante il salvataggio del plot su Excel: {e}")

    def metrics_plot(self, metrics):
        """
        Mostra (e salva su Excel) i grafici delle metriche (boxplot e line plot).
        """
        labels = list(metrics.keys())
        metric_scores = [vals if isinstance(vals, list) else [vals] for vals in metrics.values()]

        # Boxplot delle metriche
        plt.figure(figsize=(12, 6))
        # Usa "tick_labels" per evitare warning in Matplotlib >= 3.9
        plt.boxplot(metric_scores, tick_labels=labels)
        plt.xlabel('Metriche')
        plt.ylabel('Valori')
        plt.title('Metriche di validazione (Boxplot)')
        plt.grid(True)

        temp_file1 = 'temp_plot_boxplot.png'
        plt.savefig(temp_file1)
        self.save_plot_to_excel(temp_file1, 'Plot Andamento Metriche')

        # Line plot delle metriche
        num_experiments = len(metric_scores[0])  # ipotizza che tutte le liste abbiano la stessa lunghezza
        x_experiments = range(1, num_experiments + 1)

        plt.figure(figsize=(12, 6))
        for label, values in zip(labels, metric_scores):
            plt.plot(x_experiments, values, marker='o', label=label)

        plt.xlabel('Esperimenti (K)')
        plt.ylabel('Valore della metrica')
        plt.title('Andamento delle metriche nei vari esperimenti')
        plt.legend()
        plt.grid(True)

        temp_file2 = 'temp_plot_line.png'
        plt.savefig(temp_file2)
        self.save_plot_to_excel(temp_file2, 'Andamento Metriche in K')

        plt.show()

    def plot_confusion_matrix(self, y_true, y_pred):
        """
        Plot della Confusion Matrix e salvataggio su Excel.
        """
        unique_labels = [2, 4]
        cm = np.zeros((len(unique_labels), len(unique_labels)), dtype=int)
        label_to_index = {label: idx for idx, label in enumerate(unique_labels)}

        for true, pred in zip(y_true, y_pred):
            cm[label_to_index[true], label_to_index[pred]] += 1

        plt.figure(figsize=(8, 6))
        plt.imshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
        plt.title('Confusion Matrix')
        plt.colorbar()
        tick_marks = np.arange(len(unique_labels))
        plt.xticks(tick_marks, unique_labels, rotation=45)
        plt.yticks(tick_marks, unique_labels)

        fmt = 'd'
        thresh = cm.max() / 2.
        for i, j in np.ndindex(cm.shape):
            plt.text(j, i, format(cm[i, j], fmt),
                     ha="center", va="center",
                     color="white" if cm[i, j] > thresh else "black")

        plt.ylabel('True Label')
        plt.xlabel('Predicted Label')
        plt.tight_layout()

        temp_file = 'temp_confusion_matrix.png'
        plt.savefig(temp_file)
        self.save_plot_to_excel(temp_file, 'Confusion Matrix')

        plt.show()

    def plot_roc_curve(self, y_true, y_score):
        """
        Plot della ROC Curve e calcolo dell'AUC, salvataggio su Excel.
        """
        fpr = []
        tpr = []
        thresholds = np.linspace(0, 1, 100)
        for threshold in thresholds:
            tp = fp = tn = fn = 0
            for true, score in zip(y_true, y_score):
                if score >= threshold:
                    if true == 4:
                        tp += 1
                    else:
                        fp += 1
                else:
                    if true == 4:
                        fn += 1
                    else:
                        tn += 1

            fpr_val = fp / (fp + tn) if (fp + tn) else 0
            tpr_val = tp / (tp + fn) if (tp + fn) else 0
            fpr.append(fpr_val)
            tpr.append(tpr_val)

        roc_auc = np.trapezoid(tpr, fpr)  # Da NumPy >= 1.21

        plt.figure(figsize=(8, 6))
        plt.plot(fpr, tpr, color='darkorange', lw=2,
                 label=f'ROC curve (AUC = {roc_auc:.2f})')
        plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.05])
        plt.xlabel('False Positive Rate')
        plt.ylabel('True Positive Rate')
        plt.title('Receiver Operating Characteristic (ROC) Curve')
        plt.legend(loc="lower right")

        temp_file = 'temp_roc_curve.png'
        plt.savefig(temp_file)
        self.save_plot_to_excel(temp_file, 'ROC Curve')

        plt.show()

# ---------------------------------------------------------------
# Esempio di utilizzo
# ---------------------------------------------------------------
if __name__ == "__main__":
    # Chiedi all'utente soltanto il nome del file .xlsx
    filename = input("Inserisci il NOME (senza path) del file Excel (default: 'risultati_metriche.xlsx'): ").strip()
    if not filename:
        filename = "risultati_metriche.xlsx"
    elif not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    # Crea l'istanza con il solo nome file
    metrics_saver = MetricsSaver(filename=filename)

    # Esempio di dizionario di metriche
    metrics_by_experiment = {
        "Accuracy Rate": [0.8, 0.85, 0.9],
        "Error Rate": [0.2, 0.15, 0.1],
        "Sensitivity": [0.75, 0.8, 0.85],
        "Specificity": [0.85, 0.9, 0.95],
        "Geometric Mean": [0.8, 0.85, 0.9],
        "Area Under the Curve": [0.9, 0.92, 0.95]
    }

    # Salva metriche
    metrics_saver.save_metrics(metrics_by_experiment)

    # Plot (boxplot, line plot)
    metrics_saver.metrics_plot(metrics_by_experiment)

    # Esempi di y_true, y_pred, y_score
    y_true = [2, 4, 4, 2, 4, 2, 4, 4, 2, 2]
    y_pred = [2, 4, 2, 2, 4, 2, 4, 4, 2, 4]
    y_score = [0.1, 0.9, 0.4, 0.2, 0.8, 0.3, 0.7, 0.9, 0.1, 0.6]

    # Plot Confusion Matrix
    metrics_saver.plot_confusion_matrix(y_true, y_pred)

    # Plot ROC Curve
    metrics_saver.plot_roc_curve(y_true, y_score)

    # Mostra dove è stato salvato
    final_path = os.path.join(os.getcwd(), filename)
    print(f"\nTutto salvato in: {final_path}")
